"""
projections_validation_builder.py
Genera projections_validation_accounting.json: copia de baseline con líneas P&L del WIP
run rate aplicadas para todo el período de proyección (todos los meses no cerrados).

El período de proyección se toma de config.py: RUNRATE_MONTHS | FORECAST_MONTHS.
Cuando un mes se cierra contablemente, actualizar config.py y re-correr el script.

Las fuentes son los archivos "Modelo Forecast" de la carpeta W33, solapas EPM:
  API - Modelo Forecast.xlsx  |  HTML - Modelo Forecast.xlsx  |  WLs - Modelo Forecast.xlsx

Uso:
    python projections_validation_builder.py
        --wip-folder "<ruta al folder W33>"
        [--no-upload]
        [--baseline-json <ruta local>]  # evita descargar de Drive
"""
import os
import sys
import io
import json
import argparse
import pandas as pd
import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import config
import pnl_common
from pnl_common import homolog, to_num, load_glosario

DIR             = os.path.dirname(os.path.abspath(__file__))
DRIVE_FOLDER_ID = config.DRIVE_FOLDER_ID
BASELINE_NAME   = "baseline_actuals+projections.json"
OUTPUT_NAME     = "projections_validation_accounting.json"
FY_PREV         = config.CURRENT_FY - 1

COLS_OUT = ["LoB", "Canal", "Pais", "Producto",
            "P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View",
            "Fecha", "Monto USD"]
GROUP = COLS_OUT[:-1]

LOBS_KEEP = {"b2b", "b2b2c", "b2c"}

WIP_FILES = [
    "API - Modelo Forecast.xlsx",
    "HTML - Modelo Forecast.xlsx",
    "WLs - Modelo Forecast.xlsx",
]

# Período de proyección: meses no cerrados contablemente.
# Cuando un mes se cierra, reducir RUNRATE_MONTHS en config.py y re-correr.
PROJECTION_MONTHS = config.RUNRATE_MONTHS | config.FORECAST_MONTHS

# ── Concepto normalization ────────────────────────────────────────────────────
CONCEPTO_TO_BASE = {
    "orders":                     "Orders",
    "gross_bookings":             "Gross Bookings",
    "gross bookings":             "Gross Bookings",
    "up_front_incentives":        "Up Front Incentives",
    "up front incentives":        "Up Front Incentives",
    "fees":                       "Fees",
    "commercial_discounts":       "Commercial Discounts",
    "commercial discounts":       "Commercial Discounts",
    "cancellations":              "Cancellations",
    "cost_of_installments":       "Cost of Installments",
    "cost of installments":       "Cost of Installments",
    "cost_of_sales_as_principal": "Cost of Sales as Principal",
    "cost of sales as principal":  "Cost of Sales as Principal",
    "credit_card_processing":     "Credit Card Processing",
    "credit card processing":     "Credit Card Processing",
    "affiliates":                 "Affiliates",
    "white_labels_api":           "White Labels / API",
    "white labels / api":         "White Labels / API",
}

VIAJE_SPLIT_BASES = {"Orders", "Gross Bookings"}

VIAJE_NORM = {
    "domestic":      "Domestic",
    "international": "International",
    "nac":           "Domestic",
    "int":           "International",
    "dom.":          "Domestic",
    "int.":          "International",
}

TARGET_N1_OVERRIDE = {
    "orders", "orders domestic", "orders international",
    "gross bookings", "gross bookings domestic", "gross bookings international",
    "up front incentives",
    "fees",
    "commercial discounts",
    "cancellations",
    "cost of sales as principal",
    "cost of installments",
    "credit card processing",
    "affiliates",
    "white labels / api",
}

MONTHS_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


# ── EPM sheet parser ──────────────────────────────────────────────────────────

def _infer_year(month_num):
    return FY_PREV if month_num >= 4 else config.CURRENT_FY


def find_epm_sheet(wb):
    for name in wb.sheetnames:
        if "EPM" in name.upper():
            return wb[name]
    return None


def parse_epm_sheet(ws, projection_months):
    """
    Read one EPM sheet; return raw records for all dates in projection_months.
    Each record includes a 'fecha' key with the date string.
    Handles different column layouts across API / HTML / WLs models.
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Find header row ───────────────────────────────────────────────────────
    HDR_TOKENS = {"concepto", "pnl", "lob-canal", "lob", "pais", "producto", "viaje"}
    hdr_idx = None
    for i, row in enumerate(all_rows):
        if not row:
            continue
        tokens = {str(v).strip().lower() for v in row if v is not None}
        if sum(1 for t in HDR_TOKENS if t in tokens) >= 4:
            hdr_idx = i
            break
    if hdr_idx is None:
        print("  WARNING: no se encontró fila de encabezado en la solapa EPM")
        return []

    hdr = [str(v).strip().lower() if v is not None else None for v in all_rows[hdr_idx]]

    # ── Detect year/month rows above header → col → date mapping ─────────────
    year_by_col  = {}
    month_by_col = {}

    for row in all_rows[:hdr_idx]:
        if not row:
            continue
        is_year  = True
        is_month = True
        has_num  = False
        for v in row:
            if v is None:
                continue
            try:
                fv = float(v)
                has_num = True
                if not (2020 <= fv <= 2035):
                    is_year = False
                if not (1 <= fv <= 12):
                    is_month = False
            except (TypeError, ValueError):
                is_year = False
                is_month = False
        if not has_num:
            continue
        if is_year and not is_month:
            for j, v in enumerate(row):
                if v is not None:
                    try:
                        year_by_col[j] = int(float(v))
                    except Exception:
                        pass
        elif is_month:
            for j, v in enumerate(row):
                if v is not None:
                    try:
                        month_by_col[j] = int(float(v))
                    except Exception:
                        pass

    col_to_date = {}
    for j, m in month_by_col.items():
        y = year_by_col.get(j, _infer_year(m))
        col_to_date[j] = f"{y}-{m:02d}-01"

    # Also check header row for Spanish month names (fallback — only when year/month rows didn't set the column)
    for j, h in enumerate(hdr):
        if h and h in MONTHS_ES and j not in col_to_date:
            m = MONTHS_ES[h]
            y = year_by_col.get(j, _infer_year(m))
            col_to_date[j] = f"{y}-{m:02d}-01"

    # Keep only columns whose date is in projection_months
    target_cols = {j: d for j, d in col_to_date.items() if d in projection_months}

    if not target_cols:
        dates_found = sorted(set(col_to_date.values()))
        print(f"  WARNING: ninguna fecha del rango de proyección encontrada en la solapa EPM")
        print(f"           Fechas en el EPM: {dates_found}")
        print(f"           Rango esperado  : {sorted(projection_months)}")
        return []

    # ── Locate dimension columns ──────────────────────────────────────────────
    def _find(names):
        for name in names:
            idx = next((j for j, h in enumerate(hdr) if h == name), None)
            if idx is not None:
                return idx
        return None

    concepto_col  = _find(["concepto", "pnl"])
    viaje_col     = _find(["viaje"])
    pais_col      = _find(["pais"])
    producto_col  = _find(["producto"])
    lob_canal_col = _find(["lob-canal"])
    lob_col       = next((j for j, h in enumerate(hdr) if h == "lob"), None)
    canal_col     = _find(["canal"])

    def _cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # ── Read data rows ────────────────────────────────────────────────────────
    recs = []
    for row in all_rows[hdr_idx + 1:]:
        if not row or all(v is None for v in row):
            continue

        concepto_raw = _cell(row, concepto_col)
        if concepto_raw is None:
            continue
        concepto_key = str(concepto_raw).strip().lower()
        if concepto_key not in CONCEPTO_TO_BASE:
            continue

        if lob_col is not None and canal_col is not None:
            lv = str(_cell(row, lob_col) or "").strip()
            cv = str(_cell(row, canal_col) or "").strip()
            lob_canal_raw = f"{lv}-{cv}" if lv and cv else lv
        elif lob_canal_col is not None:
            lob_canal_raw = _cell(row, lob_canal_col)
        else:
            lob_canal_raw = str(_cell(row, lob_col) or "").strip()

        dim = {
            "concepto":  concepto_key,
            "viaje":     str(_cell(row, viaje_col) or "").strip(),
            "pais":      str(_cell(row, pais_col) or "").strip(),
            "producto":  str(_cell(row, producto_col) or "").strip(),
            "lob_canal": str(lob_canal_raw or "").strip(),
        }

        for col, fecha in target_cols.items():
            raw_val = _cell(row, col)
            amount = to_num(str(raw_val)) if raw_val is not None else None
            if not amount:
                continue
            recs.append({**dim, "fecha": fecha, "monto": amount})

    print(f"    {len(recs)} filas con conceptos target ({len(target_cols)} meses)")
    return recs


# ── Homologation ──────────────────────────────────────────────────────────────

def _homolog_lob(g, lob_canal_raw):
    raw_norm = "B2B-MAY" if str(lob_canal_raw).strip().lower() == "api" else lob_canal_raw
    lob   = homolog(g["lob"], raw_norm, 0, None)
    canal = homolog(g["lob"], raw_norm, 1, None)
    if lob is None and raw_norm:
        parts = str(raw_norm).split("-", 1)
        lob   = parts[0]
        canal = parts[1] if len(parts) > 1 else "Sin Canal"
    return lob, canal


def _build_linea_pnl(concepto_key, viaje_raw):
    base  = CONCEPTO_TO_BASE.get(concepto_key, concepto_key)
    viaje = VIAJE_NORM.get(viaje_raw.lower(), None)
    if base in VIAJE_SPLIT_BASES and viaje:
        return f"{base} {viaje}"
    return base


def homologate_records(recs, g):
    rows = []
    for r in recs:
        linea_pnl = _build_linea_pnl(r["concepto"], r["viaje"])
        lob, canal = _homolog_lob(g, r["lob_canal"])
        if lob is None:
            continue

        lob_l   = str(lob).lower()
        canal_l = "total" if lob_l != "b2b" else (str(canal).lower() if canal else "total")

        if lob_l not in LOBS_KEEP:
            continue

        pais = homolog(g["paises"],   r["pais"],     0, r["pais"])     or r["pais"]
        prod = homolog(g["producto"], r["producto"], 0, r["producto"]) or r["producto"]

        pnl = [homolog(g["linea"], linea_pnl, i, linea_pnl if i == 0 else None)
               for i in range(7)]

        pais_l = str(pais).lower()
        n1_l   = str(pnl[0]).lower() if pnl[0] else linea_pnl.lower()

        if pais_l == "rg" and n1_l == "cost of sales as principal":
            continue

        rows.append({
            "LoB":                 lob_l,
            "Canal":               canal_l,
            "Pais":                pais_l,
            "Producto":            str(prod).lower(),
            "P&L N1":              n1_l,
            "P&L N2":              str(pnl[1]).lower() if pnl[1] else "",
            "P&L N3":              str(pnl[2]).lower() if pnl[2] else "",
            "P&L N4":              str(pnl[3]).lower() if pnl[3] else "",
            "P&L N5":              str(pnl[4]).lower() if pnl[4] else "",
            "P&L N6":              str(pnl[5]).lower() if pnl[5] else "",
            "P&L Managerial View": str(pnl[6]).lower() if pnl[6] else "",
            "Fecha":               r["fecha"],
            "Monto USD":           r["monto"],
        })
    return rows


# ── Main data loading ─────────────────────────────────────────────────────────

def load_wip_data(wip_folder, projection_months):
    g = load_glosario()
    all_recs = []

    for fname in WIP_FILES:
        path = os.path.join(wip_folder, fname)
        if not os.path.isfile(path):
            print(f"  AVISO: no encontré {fname}")
            continue
        print(f"\n  Leyendo {fname}...")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = find_epm_sheet(wb)
        if ws is None:
            print(f"  AVISO: no hay solapa EPM en {fname}")
            wb.close()
            continue
        recs = parse_epm_sheet(ws, projection_months)
        wb.close()
        all_recs.extend(recs)

    if not all_recs:
        sys.exit("No se encontraron datos en las solapas EPM.")

    rows = homologate_records(all_recs, g)
    if not rows:
        sys.exit("Sin filas después de homologar.")

    df = pd.DataFrame(rows)
    df["Monto USD"] = pd.to_numeric(df["Monto USD"], errors="coerce").fillna(0).round(2)
    df = df.groupby(GROUP, as_index=False, dropna=False)["Monto USD"].sum()
    df = df[df["Monto USD"] != 0]
    df = df[df["P&L N1"].isin(TARGET_N1_OVERRIDE)]

    print(f"\n  Datos modelo: {len(df):,} filas para {len(df['Fecha'].unique())} meses")
    for fecha, gdf in df.groupby("Fecha"):
        print(f"    {fecha}:")
        for n1, tot in gdf.groupby("P&L N1")["Monto USD"].sum().sort_index().items():
            print(f"      {n1}: {tot:,.0f}")

    return df[COLS_OUT]


# ── Drive helpers ─────────────────────────────────────────────────────────────

def _download_baseline(svc):
    from googleapiclient.http import MediaIoBaseDownload
    q = f"name='{BASELINE_NAME}' and '{DRIVE_FOLDER_ID}' in parents and trashed=false"
    files = svc.files().list(q=q, fields="files(id)").execute().get("files", [])
    if not files:
        sys.exit(f"No encontré {BASELINE_NAME} en Drive.")
    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, svc.files().get_media(fileId=files[0]["id"]),
                               chunksize=8 * 1024 * 1024)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return json.load(buf)


def _upload(local_path, svc):
    from googleapiclient.http import MediaFileUpload
    media = MediaFileUpload(local_path, mimetype="application/json", resumable=True)
    q = f"name='{OUTPUT_NAME}' and '{DRIVE_FOLDER_ID}' in parents and trashed=false"
    ex = svc.files().list(q=q, fields="files(id)").execute().get("files", [])
    if ex:
        res = svc.files().update(fileId=ex[0]["id"], media_body=media,
                                  fields="id,size,modifiedTime").execute()
        print(f"  [Drive] actualizado {OUTPUT_NAME}: {res}")
    else:
        res = svc.files().create(body={"name": OUTPUT_NAME, "parents": [DRIVE_FOLDER_ID]},
                                  media_body=media, fields="id").execute()
        print(f"  [Drive] creado {OUTPUT_NAME}: {res}")


# ── Build ─────────────────────────────────────────────────────────────────────

def build(wip_folder, projection_months, upload, baseline_json_path=None):
    print(f"=== projections_validation_accounting ===")
    print(f"  WIP folder         : {wip_folder}")
    print(f"  Período proyección : {sorted(projection_months)}")

    svc = None

    if baseline_json_path and os.path.isfile(baseline_json_path):
        print(f"  Baseline           : {baseline_json_path} (local)")
        with open(baseline_json_path, encoding="utf-8") as f:
            baseline = json.load(f)
    else:
        print(f"  Baseline           : descargando {BASELINE_NAME} de Drive...")
        svc = pnl_common.get_drive_service()
        baseline = _download_baseline(svc)

    cols = baseline["cols"]
    rows = baseline["rows"]
    FI   = cols.index("Fecha")
    N1I  = cols.index("P&L N1")
    print(f"  Baseline           : {len(rows):,} filas")

    # Load WIP model data for all projection months
    model_df = load_wip_data(wip_folder, projection_months)
    new_rows = model_df.values.tolist()

    # Override: remove baseline rows for all projection months × target P&L N1
    removed = [r for r in rows if r[FI] in projection_months
               and str(r[N1I]).lower() in TARGET_N1_OVERRIDE]
    kept    = [r for r in rows if not (r[FI] in projection_months
               and str(r[N1I]).lower() in TARGET_N1_OVERRIDE)]

    print(f"\n  Baseline rows eliminadas ({len(projection_months)} meses): {len(removed):,}")
    print(f"  Baseline rows conservadas                              : {len(kept):,}")
    print(f"  Filas del modelo añadidas                              : {len(new_rows):,}")

    all_rows = kept + new_rows

    payload = {
        "meta": {
            "concepto":           "projections_validation_accounting",
            "projection_months":  sorted(projection_months),
            "wip_folder":         os.path.basename(wip_folder),
            "filas":              len(all_rows),
            "fechas":             sorted({r[FI] for r in all_rows}),
        },
        "cols": COLS_OUT,
        "rows": all_rows,
    }

    outdir = os.path.join(DIR, "_projections_out")
    os.makedirs(outdir, exist_ok=True)
    local = os.path.join(outdir, OUTPUT_NAME)
    with open(local, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"), default=str)
    print(f"\n  local -> {local}  ({os.path.getsize(local)/1e6:.2f} MB, {len(all_rows):,} filas)")

    if upload:
        if svc is None:
            svc = pnl_common.get_drive_service()
        _upload(local, svc)
    else:
        print("  (--no-upload: no se subió a Drive)")

    return payload


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--wip-folder", required=True,
                    help="Ruta al folder del WIP run rate (ej: '.../Run Rate/2026.08.18 - W33')")
    ap.add_argument("--no-upload", action="store_true")
    ap.add_argument("--baseline-json",
                    help="Ruta local al baseline JSON (evita descargarlo de Drive)")
    a = ap.parse_args()

    build(a.wip_folder, PROJECTION_MONTHS, not a.no_upload, a.baseline_json)
