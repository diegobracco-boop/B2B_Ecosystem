"""
pnl_common.py — Rutas y autenticación compartidas para los uploaders del P&L.
Portable multi-usuario: resuelve la carpeta OneDrive sin importar el usuario,
auto-detecta el forecast del mes más reciente, y autentica a Drive con scope
COMPLETO (credentials_drive.json + token_drive.json) para que cualquiera del
equipo pueda sobrescribir los JSON en Drive.
"""
import os
import re
import json
import datetime
from pathlib import Path

_DIR = os.path.dirname(os.path.abspath(__file__))
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]


# ── Rutas portables ──────────────────────────────────────────────────────────
def _onedrive_root():
    """Raíz de OneDrive corporativo del usuario (no hardcodeada)."""
    od = os.environ.get("OneDriveCommercial")
    if not od:
        home = os.path.expanduser("~")
        # probar nombres conocidos de la carpeta corporativa Despegar
        for candidate in ("despegar365", "OneDrive - despegar365", "OneDrive - Despegar365"):
            p = os.path.join(home, candidate)
            if os.path.isdir(p):
                od = p
                break
    if not od:
        od = os.environ.get("OneDrive") or os.path.join(os.path.expanduser("~"), "OneDrive")
    return od


def get_base_dir():
    """Carpeta OneDrive compartida del equipo, resuelta por usuario (no hardcodeada)."""
    return os.path.join(_onedrive_root(), "Control de Gestión - 2026-27", "B2B & WLs")


def get_pbi_inputs_dir():
    """Carpeta 'Planning-PBI - Inputs Power Bi' (bases raw de proyecciones/actuals)."""
    return os.path.join(_onedrive_root(), "Planning-PBI - Inputs Power Bi")


def get_glosario_path():
    return os.path.join(get_base_dir(), "Proyectos IA", "BITUBIA", "Glosario.xlsx")


def get_reverso_axi_path():
    return os.path.join(get_pbi_inputs_dir(), "Actuals", "Reverso AxI.xlsx")


def get_toqan_dir():
    return os.path.join(get_base_dir(), "Proyectos IA", "BITUBIA", "Output Toqan")


def get_revision_dir():
    return os.path.join(get_base_dir(), "Proyectos IA", "Codigo - revision P&L")


# Versión del modelo Forecast vigente. ACTUALIZAR A MANO cuando el equipo confirme
# uno nuevo. NO auto-detectar la carpeta más reciente: puede haber borradores/olvidados
# (ej. 2026.07.21 quedó sin usar). El vigente es el 2026.07.14.
FORECAST_VERSION = "2026.07.14"

def get_models_dir():
    """Subcarpeta Forecast del modelo VIGENTE (fija, ver FORECAST_VERSION)."""
    return os.path.join(get_base_dir(), "Forecast", FORECAST_VERSION)


# ── Auth Drive (scope completo, compartible por el equipo) ────────────────────
def get_drive_service():
    """
    Cliente de Drive con scope 'drive' completo desde credentials_drive.json +
    token_drive.json (junto a este archivo). Permite sobrescribir archivos creados
    por otros usuarios (necesario para que corra todo el equipo).
    Fallback: token de clasp (~/.clasprc.json, scope drive.file — solo archivos propios).
    """
    from googleapiclient.discovery import build
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request

    creds_file = os.path.join(_DIR, "credentials_drive.json")
    token_file = os.path.join(_DIR, "token_drive.json")

    if os.path.exists(creds_file):
        from google_auth_oauthlib.flow import InstalledAppFlow
        creds = None
        if os.path.exists(token_file):
            creds = Credentials.from_authorized_user_file(token_file, DRIVE_SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(creds_file, DRIVE_SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_file, "w", encoding="utf-8") as f:
                f.write(creds.to_json())
        return build("drive", "v3", credentials=creds)

    # Fallback: clasp token (scope drive.file)
    clasprc = Path.home() / ".clasprc.json"
    if not clasprc.exists():
        raise FileNotFoundError(
            "Falta credentials_drive.json (recomendado) o ~/.clasprc.json. "
            "Correr auth_drive.py una vez, o 'clasp login'."
        )
    tok = json.loads(clasprc.read_text())["tokens"]["default"]
    creds = Credentials(
        token=tok["access_token"], refresh_token=tok["refresh_token"],
        token_uri="https://oauth2.googleapis.com/token",
        client_id=tok["client_id"], client_secret=tok["client_secret"],
        expiry=datetime.datetime.utcfromtimestamp(tok["expiry_date"] / 1000),
    )
    if not creds.valid:
        creds.refresh(Request())
    return build("drive", "v3", credentials=creds)


# ── Constantes compartidas ────────────────────────────────────────────────────
MMM = {"jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
       "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12"}

DIM_COLS = ["Marca", "LoB", "Canal", "Pais", "Producto",
            "P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]
OUT_COLS = DIM_COLS + ["Fecha", "Monto USD"]
LOBS_KEEP = ["b2b", "b2b2c", "b2c"]

EXCL_AXI_MARCA = {"koin", "intercompany", "dfinance"}
PAIS_AXI_MAP = {
    "brasil": "brasil", "argentina": "argentina", "mexico": "mexico", "colombia": "colombia",
    "chile": "chile", "peru": "peru", "ecuador": "ecuador",
    "others": "others countries", "others countries": "others countries",
    "ops + rg": "others countries",
}


# ── Helpers compartidos ───────────────────────────────────────────────────────
def homolog(m, val, idx=0, default=None):
    if val is None:
        return default
    r = m.get(str(val).strip().lower())
    if not r:
        return default
    return r[idx] if idx < len(r) else default


def to_num(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() == "#missing":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def lower_all(df):
    for c in df.columns:
        if c != "Monto USD":
            df[c] = df[c].astype(str).str.lower()
    return df


def load_glosario():
    import openpyxl
    wb = openpyxl.load_workbook(get_glosario_path(), read_only=True, data_only=True)

    def sheet_map(name, key_col, val_cols):
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        m = {}
        for r in rows:
            if r is None or r[key_col] is None:
                continue
            k = str(r[key_col]).strip().lower()
            m[k] = [(r[c] if c < len(r) and r[c] is not None else None) for c in val_cols]
        return m

    g = {
        "marca":    sheet_map("Marca",    0, [1]),
        "paises":   sheet_map("Paises",   0, [1]),
        "producto": sheet_map("Producto", 0, [1]),
        "lob":      sheet_map("LOB",      0, [1, 2]),
        "linea":    sheet_map("Linea P&L", 0, [1, 2, 3, 4, 5, 6, 7]),
    }
    wb.close()
    return g


def add_axi(plana, axi):
    import pandas as pd
    if axi is None or axi.empty:
        return plana
    combined = pd.concat([plana, axi[OUT_COLS]], ignore_index=True)
    return combined.groupby(DIM_COLS + ["Fecha"], as_index=False, dropna=False)["Monto USD"].sum()


def load_axi_sheet(g, sheet, keep_dates=None):
    """Lee una hoja del Reverso AxI, homologa y devuelve DataFrame en formato OUT_COLS."""
    import openpyxl
    import pandas as pd
    _NIVELES = ["P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]
    wb = openpyxl.load_workbook(get_reverso_axi_path(), read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    idx = {str(h).strip().lower(): j for j, h in enumerate(hdr)}
    date_cols = {}
    for j, h in enumerate(hdr):
        hs = str(h).strip().lower()
        if len(hs) == 6 and hs[:3] in MMM and hs[3] == "-":
            fecha = f"20{hs[4:]}-{MMM[hs[:3]]}-01"
            if keep_dates is None or fecha in keep_dates:
                date_cols[j] = fecha
    recs = []
    for r in rows:
        if r is None:
            continue
        marca_raw = str(r[idx["brand"]]).strip().lower() if r[idx["brand"]] is not None else ""
        if marca_raw in EXCL_AXI_MARCA:
            continue
        base = {
            "Marca":    homolog(g["marca"], r[idx["brand"]], 0, r[idx["brand"]]),
            "Producto": homolog(g["producto"], r[idx["product"]], 0, r[idx["product"]]),
            "Pais":     PAIS_AXI_MAP.get(str(r[idx["country"]]).strip().lower(),
                                         str(r[idx["country"]]).strip().lower()) if r[idx["country"]] is not None else None,
            "LoB":      homolog(g["lob"], r[idx["lob"]], 0, None),
            "Canal":    homolog(g["lob"], r[idx["lob"]], 1, None),
        }
        if base["LoB"] is None and r[idx["lob"]] is not None:
            sp = str(r[idx["lob"]]).split("-", 1)
            base["LoB"], base["Canal"] = sp[0], (sp[1] if len(sp) > 1 else "Sin Canal")
        linea = r[idx["linea p&l"]]
        for i, col in enumerate(_NIVELES):
            base[col] = homolog(g["linea"], linea, i, linea)
        for j, fecha in date_cols.items():
            val = to_num(r[j]) if j < len(r) else None
            if val:
                rec = dict(base); rec["Fecha"] = fecha; rec["Monto USD"] = val
                recs.append(rec)
    wb.close()
    if not recs:
        return pd.DataFrame(columns=OUT_COLS)
    axi = pd.DataFrame(recs)
    axi = axi.groupby(DIM_COLS + ["Fecha"], as_index=False, dropna=False)["Monto USD"].sum()
    return lower_all(axi)
