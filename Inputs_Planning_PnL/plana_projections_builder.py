"""
plana_projections_builder.py
Genera las PLANAS de proyecciones (Budget / Forecast / Run Rate / Last Run Rate)
desde los CSV de Planning-PBI, replicando el proceso del agente interno (Partes 3-6).

Uso:
    python plana_projections_builder.py <base> [--con-ppa]
    <base> ∈ budget | forecast | runrate | lastrunrate

Default = SIN PPA  -> suma el efecto de 'Reverso AxI.xlsx' (hoja Budget).
--con-ppa           -> NO suma el reverso (procesa normal).

Output: plana_<base>.csv  (long: Marca|LoB|Canal|Pais|Producto|P&L N1..N6|P&L Managerial View|Fecha|Monto USD)
"""
import os
import sys
import pandas as pd
import pnl_common

# Consola Windows (cp1252) → forzar UTF-8 para no crashear con acentos/flechas
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ── Meses del año fiscal FY27 (Abr 2026 → Mar 2027) ──────────────────────────
# OJO: el spec del agente listaba 2025-04..2026-03; se usa 2026-04..2027-03 (FY27)
# para alinear con la hoja 'Budget' del Reverso AxI (apr-26..mar-27) y con las landings.
FISCAL_DATES = [
    "2026-04-01", "2026-05-01", "2026-06-01", "2026-07-01", "2026-08-01", "2026-09-01",
    "2026-10-01", "2026-11-01", "2026-12-01", "2027-01-01", "2027-02-01", "2027-03-01",
]
FORECAST_DROP_DATES = {"2026-04-01", "2026-05-01", "2026-06-01", "2026-07-01"}  # solo Forecast

PBI = pnl_common.get_pbi_inputs_dir()
BASES = {
    "budget":      {"files": ["Budget/Budget 2027 - Legal Entity ALL.csv",
                              "Budget/Budget 2027 - Legal Entity NA.csv"],       "is_forecast": False},
    "forecast":    {"files": ["Forecast/FQ1 - Legal Entity ALL.csv",
                              "Forecast/FQ1 - Legal Entity NA.csv"],             "is_forecast": True},
    "forecast_v2": {"files": ["Forecast/FQ1 - Legal Entity ALL V2.csv",
                              "Forecast/FQ1 - Legal Entity NA V2.csv"],          "is_forecast": True},
    "runrate":     {"files": ["Run Rate/RR - Legal Entity ALL - 27.csv",
                              "Run Rate/RR - Legal Entity NA - 27.csv"],         "is_forecast": False},
    "lastrunrate": {"files": ["Run Rate/LRR - Legal Entity ALL.csv",
                              "Run Rate/LRR - Legal Entity NA.csv"],             "is_forecast": False},
}

DIM_COLS = ["Marca", "LoB", "Canal", "Pais", "Producto",
            "P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]
OUT_COLS = DIM_COLS + ["Fecha", "Monto USD"]
LOBS_KEEP = ["b2b", "b2b2c", "b2c"]   # LOBs del dashboard; el resto (dmc/stays/koin/...) se descarta

# ── Filtros de exclusión (Parte 3.6) ─────────────────────────────────────────
EXCL_MARCA   = {"br_4001", "interco"}
EXCL_LOBCANAL = {"financial services business", "lob_00"}
EXCL_LINEA = {
    "3rd parties fx", "cash & equivalents fx", "factoring expense", "financial expense",
    "financial income", "income tax", "other adjustments non-fx", "other fx", "withholdings tax",
    "intercompany fx", "ndf fixing", "bank charges", "tax penalties", "guarantee charges",
    "neteo fx", "other income/expense non-fx", "other interest expense", "other interest income",
}
# Exclusiones extra de la fuente AXI (Parte 4.2)
EXCL_AXI_MARCA = {"koin", "intercompany", "dfinance"}

PAIS_AXI_MAP = {
    "brasil": "brasil", "argentina": "argentina", "mexico": "mexico", "colombia": "colombia",
    "chile": "chile", "peru": "peru", "ecuador": "ecuador",
    "others": "others countries", "others countries": "others countries",
    "ops + rg": "others countries",
}


# ── Glosario ─────────────────────────────────────────────────────────────────
def load_glosario():
    import openpyxl
    wb = openpyxl.load_workbook(pnl_common.get_glosario_path(), read_only=True, data_only=True)

    def sheet_map(name, key_col, val_cols):
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)  # header
        m = {}
        for r in rows:
            if r is None or r[key_col] is None:
                continue
            k = str(r[key_col]).strip().lower()
            m[k] = [(r[c] if c < len(r) and r[c] is not None else None) for c in val_cols]
        return m

    g = {
        "marca":    sheet_map("Marca",    0, [1]),
        "paises":   sheet_map("Paises",   0, [1]),
        "producto": sheet_map("Producto", 0, [1]),
        "lob":      sheet_map("LOB",      0, [1, 2]),          # -> [LOB, CANAL]
        "linea":    sheet_map("Linea P&L", 0, [1, 2, 3, 4, 5, 6, 7]),  # -> N1..N6, Managerial
    }
    wb.close()
    return g


def homolog(m, val, idx=0, default=None):
    if val is None:
        return default
    r = m.get(str(val).strip().lower())
    if not r:
        return default
    return r[idx] if idx < len(r) else default


# ── Parser especial de los CSV (Parte 3.2) ───────────────────────────────────
def parse_line(contenido):
    partes, dentro, actual = [], False, ""
    for ch in contenido:
        if ch == '"':
            dentro = not dentro
        elif ch == ',' and not dentro:
            partes.append(actual.strip()); actual = ""
        else:
            actual += ch
    if actual:
        partes.append(actual.strip())
    if len(partes) < 14:
        return None
    cuenta = partes[0]
    meses  = partes[1:13]
    meta   = partes[13].strip('"').split(',')
    return cuenta, meses, meta


def to_num(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() == "#missing":
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ── Lectura + parseo de una base (ALL + NA) ──────────────────────────────────
def read_base(files):
    recs = []
    for rel in files:
        path = os.path.join(PBI, rel)
        with open(path, encoding="utf-8-sig") as f:
            next(f, None)  # header
            for line in f:
                line = line.rstrip("\n").rstrip("\r")
                if not line:
                    continue
                # Las filas vienen envueltas en comillas (toda la fila = 1 campo, con "" internas):
                # desenvolver y desescapar antes de aplicar el parser del POV.
                if len(line) >= 2 and line[0] == '"' and line[-1] == '"':
                    line = line[1:-1].replace('""', '"')
                p = parse_line(line)
                if p is None:
                    continue
                cuenta, meses, meta = p
                if len(meta) < 8:
                    continue
                rec = {
                    "Linea P&L":  cuenta,
                    "Escenario":  meta[1] if len(meta) > 1 else None,
                    "Marca":      meta[3] if len(meta) > 3 else None,
                    "Viaje":      meta[4] if len(meta) > 4 else None,
                    "Pais":       meta[5] if len(meta) > 5 else None,
                    "Producto":   meta[6] if len(meta) > 6 else None,
                    "LOB-CANAL":  meta[7] if len(meta) > 7 else None,
                }
                for i, d in enumerate(FISCAL_DATES):
                    rec[d] = to_num(meses[i]) if i < len(meses) else None
                recs.append(rec)
    return pd.DataFrame(recs)


# ── Transform GB/Orders (Parte 3.5) ──────────────────────────────────────────
def gb_orders_transform(df):
    v = df["Viaje"].astype(str).str.strip().str.lower()
    l = df["Linea P&L"].astype(str).str.strip().str.lower()
    df.loc[(v == "d_0001") & (l == "gb"),       "Linea P&L"] = "gross bookings domestic"
    df.loc[(v == "d_0001") & (l == "orders_s"), "Linea P&L"] = "orders domestic"
    df.loc[(v == "d_0002") & (l == "gb"),       "Linea P&L"] = "gross bookings international"
    df.loc[(v == "d_0002") & (l == "orders_s"), "Linea P&L"] = "orders international"
    return df


def apply_exclusions(df, is_forecast):
    df = df[~df["Marca"].astype(str).str.strip().str.lower().isin(EXCL_MARCA)]
    df = df[~df["LOB-CANAL"].astype(str).str.strip().str.lower().isin(EXCL_LOBCANAL)]
    df = df[~df["Linea P&L"].astype(str).str.strip().str.lower().isin(EXCL_LINEA)]
    return df


def homologate(df, g):
    df = df.copy()
    df["Marca"]    = df["Marca"].apply(lambda x: homolog(g["marca"], x, 0, x))
    df["Pais"]     = df["Pais"].apply(lambda x: homolog(g["paises"], x, 0, x))
    df["Producto"] = df["Producto"].apply(lambda x: homolog(g["producto"], x, 0, x))
    df["LoB"]      = df["LOB-CANAL"].apply(lambda x: homolog(g["lob"], x, 0, None))
    df["Canal"]    = df["LOB-CANAL"].apply(lambda x: homolog(g["lob"], x, 1, None))
    # fallback split si no está en glosario
    miss = df["LoB"].isna()
    if miss.any():
        split = df.loc[miss, "LOB-CANAL"].astype(str).str.split("-", n=1, expand=True)
        df.loc[miss, "LoB"]   = split[0]
        df.loc[miss, "Canal"] = split[1] if split.shape[1] > 1 else "Sin Canal"
    for i, col in enumerate(["P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]):
        df[col] = df["Linea P&L"].apply(lambda x, i=i: homolog(g["linea"], x, i, x))
    return df


def melt_and_group(df):
    id_vars = ["Marca", "LoB", "Canal", "Pais", "Producto",
               "P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]
    long = df.melt(id_vars=id_vars, value_vars=FISCAL_DATES, var_name="Fecha", value_name="Monto USD")
    long = long.dropna(subset=["Monto USD"])
    long["Monto USD"] = pd.to_numeric(long["Monto USD"], errors="coerce")
    long = long.dropna(subset=["Monto USD"])
    long = long[long["Monto USD"] != 0]
    grouped = long.groupby(id_vars + ["Fecha"], as_index=False, dropna=False)["Monto USD"].sum()
    return grouped


def lower_all(df):
    for c in df.columns:
        if c != "Monto USD":
            df[c] = df[c].astype(str).str.lower()
    return df


# ── Parte 4: Reverso AxI (SIN PPA) ───────────────────────────────────────────
def load_axi_budget(g):
    """Hoja 'Budget' del Reverso AxI, homologada al mismo formato que la plana."""
    import openpyxl
    wb = openpyxl.load_workbook(pnl_common.get_reverso_axi_path(), read_only=True, data_only=True)
    ws = wb["Budget"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    # cols fecha: 'apr-26'.. -> YYYY-MM-DD
    mmm = {"jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
           "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12"}
    date_cols = {}
    for j, h in enumerate(hdr):
        hs = str(h).strip().lower()
        if len(hs) == 6 and hs[:3] in mmm and hs[3] == "-":
            date_cols[j] = f"20{hs[4:]}-{mmm[hs[:3]]}-01"
    idx = {str(h).strip().lower(): j for j, h in enumerate(hdr)}
    recs = []
    for r in rows:
        if r is None:
            continue
        marca_raw = str(r[idx["brand"]]).strip().lower() if r[idx["brand"]] is not None else ""
        if marca_raw in EXCL_AXI_MARCA:
            continue
        base = {
            "Marca":    homolog(g["marca"], r[idx["brand"]], 0, r[idx["brand"]]),
            "Producto": homolog(g["producto"], r[idx["product"]], 0, r[idx["product"]]),
            "Pais":     PAIS_AXI_MAP.get(str(r[idx["country"]]).strip().lower(),
                                        str(r[idx["country"]]).strip().lower()) if r[idx["country"]] is not None else None,
            "LoB":      homolog(g["lob"], r[idx["lob"]], 0, None),
            "Canal":    homolog(g["lob"], r[idx["lob"]], 1, None),
        }
        linea = r[idx["linea p&l"]]
        for i, col in enumerate(["P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]):
            base[col] = homolog(g["linea"], linea, i, linea)
        if base["LoB"] is None and r[idx["lob"]] is not None:
            sp = str(r[idx["lob"]]).split("-", 1)
            base["LoB"], base["Canal"] = sp[0], (sp[1] if len(sp) > 1 else "Sin Canal")
        for j, fecha in date_cols.items():
            val = to_num(r[j]) if j < len(r) else None
            if val:
                rec = dict(base); rec["Fecha"] = fecha; rec["Monto USD"] = val
                recs.append(rec)
    wb.close()
    axi = pd.DataFrame(recs)
    if axi.empty:
        return axi
    axi = axi.groupby(DIM_COLS + ["Fecha"], as_index=False, dropna=False)["Monto USD"].sum()
    return lower_all(axi)


def add_axi(plana, axi):
    """Suma por combinación de dimensiones (existente -> suma; nueva -> fila nueva)."""
    if axi is None or axi.empty:
        return plana
    combined = pd.concat([plana, axi[OUT_COLS]], ignore_index=True)
    return combined.groupby(DIM_COLS + ["Fecha"], as_index=False, dropna=False)["Monto USD"].sum()


# ── Verificación (Parte 5) ───────────────────────────────────────────────────
def verify(df):
    print("\n--- VERIFICACIÓN ---")
    print(f"filas: {len(df):,}")
    print("nulos ->", {c: int(df[c].isna().sum()) for c in ["Marca", "LoB", "Canal", "Pais", "Producto"]})
    print("únicos ->", {c: int(df[c].nunique()) for c in ["Marca", "Pais", "Producto", "LoB"]})
    print("Pais:", sorted(df["Pais"].dropna().unique())[:15])
    print("Producto:", sorted(df["Producto"].dropna().unique())[:15])
    print("P&L N2:", sorted(df["P&L N2"].dropna().unique())[:20])
    print("Fecha rango:", df["Fecha"].min(), "->", df["Fecha"].max())
    print("Monto USD total:", round(df["Monto USD"].sum(), 1))


# ── Main ─────────────────────────────────────────────────────────────────────
def build(base, con_ppa=False):
    cfg = BASES[base]
    print(f"=== plana {base} ({'CON' if con_ppa else 'SIN'} PPA) ===")
    df = read_base(cfg["files"])
    print(f"  leídas {len(df):,} filas (ALL+NA)")
    df = gb_orders_transform(df)
    df = apply_exclusions(df, cfg["is_forecast"])
    if cfg["is_forecast"]:
        for d in FORECAST_DROP_DATES:
            df[d] = None  # Forecast: excluir Abr-Jul (aún no proyectados)
    g = load_glosario()
    df = homologate(df, g)
    plana = melt_and_group(df)
    plana = lower_all(plana)[OUT_COLS]
    plana = plana[plana["LoB"].isin(LOBS_KEEP)]   # solo LOBs del dashboard (b2b/b2b2c/b2c)
    if not con_ppa:
        axi = load_axi_budget(g)
        if axi is not None and not axi.empty:
            axi = axi[axi["LoB"].isin(LOBS_KEEP)]
            axi = axi[axi["Fecha"].isin(set(plana["Fecha"].unique()))]   # AXI solo a meses presentes en la base
        print(f"  reverso AxI (Budget): {0 if axi is None else len(axi):,} filas a sumar")
        plana = add_axi(plana, axi)[OUT_COLS]
    verify(plana)
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"plana_{base}.csv")
    plana.to_csv(out, index=False, encoding="utf-8")
    print(f"\nOK -> {out}")
    return plana


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    con_ppa = "--con-ppa" in sys.argv
    if not args or args[0] not in BASES:
        sys.exit(f"Uso: python plana_projections_builder.py <{'|'.join(BASES)}> [--con-ppa]")
    build(args[0], con_ppa)
