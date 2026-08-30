"""
okr_builder.py
Genera okr.json combinando cuatro fuentes:

  1. KRs B2B2C Op.Contribution + B2B NR Core/New Markets (mismo input contable
     que usa la landing P&L_Accounting)
       <- baseline_actuals+projections.json  → "Run Rate/Actuals"
       <- budget.json                        → "Budget"

  2. KRs B2B2C Hunting/Farming (New / Existing Account Net Revenues):
       Definición única de "hunting" = flag New/Existing de daily_b2b2c_data.json
       (account_type en actuals, stage en budget/runrate). 2026-08-30: se descartó
       la lista HUNTING_PARTNERS (coincidía ~2% con el flag para budget).
       <- daily_b2b2c_data.json .budget (stage=='New')                 → Hunting "Budget"
       <- daily_b2b2c_data.json .actuals (account_type=='New')          → Hunting "Run Rate/Actuals", meses cerrados
       <- daily_b2b2c_data.json .runrate (stage=='New'), fallback .budget → Hunting "Run Rate/Actuals", meses proyectados
       Existing = Total B2B2C NR contable (fuente 1) − Hunting.

  3. KRs manuales del GSheet "Input_OKR" (Sign New Partnership; Air Net Revenue
     from suppliers — manual hasta definir fuente, 2026-08-25).

  4. KR "Monthly Buying Agencies" (B2B): flat xlsx agencias_okr.xlsx llenado a
     mano (el estado de la orden en el datalake muta → hay que congelar el
     snapshot mensual). Vida útil hasta Sep-2026; Oct-2026 cambia la lógica.

Mismo criterio que Dashboard_B2B_WLs/Codigo_OKR.js (computeOKR_), consolidado acá para
que la landing deje de calcular todo en vivo y solo lea este JSON.

Uso:
    python okr_builder.py
    python okr_builder.py --no-upload
"""

import os, sys, io, json, re, argparse
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import config
import pnl_common

DIR             = os.path.dirname(os.path.abspath(__file__))
OUTPUT_NAME     = "okr.json"
DRIVE_FOLDER_ID = config.DRIVE_FOLDER_ID

# ── IDs Drive ──────────────────────────────────────────────────────────────────
GESTIONAL_FILE_ID  = "1wvle0UIVZV7ocCSl8OawOfIGVz_It5kh"   # _pnl_gestional_data.json (solo para 'last_actual_ym' = corte de actuals)
BASELINE_FILE_ID   = "1Su36jhCMdNgC6nixxX5TyvtCI4ESG2Tv"   # baseline_actuals+projections.json
BUDGET_FILE_ID     = "1f2JF8pq7gtpxfdkVzbT9wvamn_ny3RBW"   # budget.json
DAILY_B2B2C_FILE_ID = "1Ukcx4e-dwCZ2VqesWwVN_1Jnt6r2AZdX"  # daily_b2b2c_data.json (transaccional)
OKR_FILE_ID        = "1cEidr8aoYgm4S7ugm05Wv-SMnz8GbtUj"   # okr.json (output)

SHEET_ID    = "1RVmTXDyyugCUXJ0f6JG_croNxWNLlOLm4eAs8F52u2c"
SHEET_RANGE = "Input_OKR"

# KR "Monthly Buying Agencies" (B2B): flat xlsx llenado a mano (periodo|escenario|valor).
# El estado de la orden en el datalake muta con el tiempo → hay que congelar el
# snapshot de cada mes. Indicador con vida hasta Sep-2026; Oct-2026 cambia la lógica.
AGENCIAS_FLAT = os.path.join(DIR, "agencias_okr.xlsx")

COLS_OUT = ["Periodo", "Escenario", "LoB", "Pais", "Producto", "KR", "Valor"]

# ── Segmentación de mercados B2B (2026-08-30 — Diego) ──
#   Core Markets = Brasil + Mexico + Others Countries (el canónico lo emite
#                  "others countries"; se aceptan ambas grafías).
#   New Markets  = TODO lo demás, incluido OPS/RG (su NR debería tender a cero).
CORE_MARKETS = {"brasil", "mexico", "other countries", "others countries"}

# KRs que vienen del GSheet Input_OKR (manuales) — el resto se calcula.
# Air Net Revenue from suppliers: manual hasta que el equipo defina de dónde sale (2026-08-25).
# Monthly Buying Agencies: NO va acá — se lee de AGENCIAS_FLAT (ver _read_agencias_flat).
MANUAL_KRS = {
    "Sign New Partnership",
    "Air Net Revenue from suppliers",
}


# ── Drive helpers ──────────────────────────────────────────────────────────────

def _download_json(svc, file_id, label):
    from googleapiclient.http import MediaIoBaseDownload
    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, svc.files().get_media(fileId=file_id),
                              chunksize=8 * 1024 * 1024)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    data = json.load(buf)
    print(f"  [Drive] {label}: OK")
    return data


def _upload(local_path, svc):
    from googleapiclient.http import MediaFileUpload
    media = MediaFileUpload(local_path, mimetype="application/json", resumable=True)
    res = svc.files().update(fileId=OKR_FILE_ID, media_body=media,
                              fields="id,size,modifiedTime").execute()
    print(f"  [Drive] actualizado {OUTPUT_NAME}: id={res['id']}  size={res.get('size','?')}")


# ── Sheets helper ──────────────────────────────────────────────────────────────

def _read_sheet(token_file):
    from googleapiclient.discovery import build
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request

    tok = json.load(open(token_file))
    creds = Credentials(
        token=tok.get("token"), refresh_token=tok.get("refresh_token"),
        token_uri=tok.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=tok.get("client_id"), client_secret=tok.get("client_secret"),
        scopes=["https://www.googleapis.com/auth/drive",
                "https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    if not creds.valid:
        creds.refresh(Request())
    sheets = build("sheets", "v4", credentials=creds)
    result = sheets.spreadsheets().values().get(
        spreadsheetId=SHEET_ID, range=SHEET_RANGE
    ).execute()
    return result.get("values", [])


# ── Gestional B2B2C ───────────────────────────────────────────────────────────

def _ym_to_periodo(ym):
    """'2026-04' → '2026-04-01'"""
    return ym + "-01"


def _daily_new_by_periodo(block, nr_field="net_revenue"):
    """daily_b2b2c_data.json bloque {cols,rows} → {periodo: NR de cuentas 'New'}.
    Los bloques budget/runrate traen el flag en la columna 'stage'."""
    cols = block.get("cols", [])
    rows = block.get("rows", [])
    if not cols or not rows:
        return {}
    try:
        iF, iS, iN = cols.index("fecha"), cols.index("stage"), cols.index(nr_field)
    except ValueError:
        return {}
    out = defaultdict(float)
    for r in rows:
        if str(r[iS]) == "New":
            out[_ym_to_periodo(str(r[iF])[:7])] += float(r[iN] or 0)
    return out


def _compute_b2b2c_hunting_farming(daily_json, nr_b2b2c_by_scen, last_actual_ym):
    """
    Devuelve rows OKR para New (Hunting) / Existing (Farming) Account Net Revenues B2B2C.

    Definición única de "hunting" = flag New/Existing de daily_b2b2c_data.json
    (account_type en .actuals, stage en .budget/.runrate). 2026-08-30: se descartó
    la lista HUNTING_PARTNERS (coincidía ~2% con el flag para budget).

      · Budget           : Hunting = daily_b2b2c.budget (stage=='New')
      · Run Rate/Actuals :
          - meses <= last_actual_ym : daily_b2b2c.actuals (account_type=='New')
          - meses  > last_actual_ym : daily_b2b2c.runrate (stage=='New'),
                                      fallback daily_b2b2c.budget si el mes no está en runrate
      · Existing = Total B2B2C NR contable (nr_b2b2c_by_scen) − Hunting, por escenario/mes.
    """
    corte = _ym_to_periodo(last_actual_ym)   # 'YYYY-MM-01'

    act_new = defaultdict(float)
    for r in daily_json.get("actuals", []):
        if str(r.get("account_type", "")) != "New":
            continue
        ym = str(r.get("fecha", ""))[:7]
        if ym:
            act_new[_ym_to_periodo(ym)] += float(r.get("net_revenues") or 0)

    bud_new = _daily_new_by_periodo(daily_json.get("budget", {}))
    rr_new  = _daily_new_by_periodo(daily_json.get("runrate", {}))

    def hunting_rr(periodo):
        if periodo <= corte:
            return act_new.get(periodo, 0.0)
        if periodo in rr_new:
            return rr_new[periodo]
        return bud_new.get(periodo, 0.0)

    HUNTING_BY_SCEN = {
        "Run Rate/Actuals": hunting_rr,
        "Budget":           lambda periodo: bud_new.get(periodo, 0.0),
    }

    rows = []
    for scen_label, hunting_fn in HUNTING_BY_SCEN.items():
        total_by_periodo = nr_b2b2c_by_scen.get(scen_label, {})
        for periodo in sorted(total_by_periodo):
            total = total_by_periodo.get(periodo)
            if total is None:
                continue   # sin total contable ese mes -> no calculamos Existing
            hunting = hunting_fn(periodo)
            farming = total - hunting
            rows.append([periodo, scen_label, "B2B2C", "Total", "Total",
                         "New Account Net Revenues", round(hunting, 2)])
            rows.append([periodo, scen_label, "B2B2C", "Total", "Total",
                         "Existing Account Net Revenues", round(farming, 2)])

    n_meses = len(set(r[0] for r in rows))
    print(f"  [Hunting/Farming] B2B2C NR: {len(rows)} filas ({n_meses} meses) · corte actuals = {last_actual_ym}")
    return rows


# ── Canonical JSON (baseline / budget) ────────────────────────────────────────

def _compute_from_canonical(canon_data, scen_label):
    """
    Calcula desde un JSON canónico (baseline o budget) — mismo input que usa la
    landing P&L_Accounting:
      - B2B2C Op. Contribution   (N5 = 'operating contribution', lob = b2b2c)
      - B2B NR Core Markets      (N3 = 'net revenue', lob = b2b, pais ∈ CORE_MARKETS)
      - B2B NR New Markets       (N3 = 'net revenue', lob = b2b, pais ∉ CORE_MARKETS — incluye OPS/RG)
    Air Net Revenue from suppliers NO se calcula acá: es manual (Input_OKR)
    hasta que el equipo defina la fuente (2026-08-25).
    Devuelve además (no como KR propio) el total B2B2C Net Revenue por mes,
    que usa _compute_b2b2c_hunting_farming para calcular "Existing".
    """
    cols = canon_data["cols"]
    iL   = cols.index("LoB")
    iN3  = cols.index("P&L N3")
    iN5  = cols.index("P&L N5")
    iP   = cols.index("Pais")
    iF   = cols.index("Fecha")
    iM   = cols.index("Monto USD")

    op_cont    = defaultdict(float)   # B2B2C Op.Contribution
    nr_core    = defaultdict(float)   # B2B NR Core Markets
    nr_new     = defaultdict(float)   # B2B NR New Markets
    nr_b2b2c   = defaultdict(float)   # B2B2C NR total (no es KR propio)

    for r in canon_data["rows"]:
        lob   = str(r[iL]).lower()
        n3    = str(r[iN3]).lower()
        n5    = str(r[iN5]).lower()
        pais  = str(r[iP]).lower()
        fecha = r[iF]
        monto = float(r[iM] or 0)

        if lob == "b2b2c" and n5 == "operating contribution":
            op_cont[fecha] += monto

        if lob == "b2b2c" and n3 == "net revenue":
            nr_b2b2c[fecha] += monto

        if lob == "b2b" and n3 == "net revenue":
            if pais in CORE_MARKETS:
                nr_core[fecha] += monto
            else:
                nr_new[fecha] += monto

    rows = []
    for fecha in sorted(set(list(op_cont) + list(nr_core) + list(nr_new))):
        if fecha in op_cont:
            rows.append([fecha, scen_label, "B2B2C", "Total", "Total",
                         "Op. Contribution", round(op_cont[fecha], 2)])
        if fecha in nr_core:
            rows.append([fecha, scen_label, "B2B", "Core Markets", "Total",
                         "Net revenues Core Markets", round(nr_core[fecha], 2)])
        if fecha in nr_new:
            rows.append([fecha, scen_label, "B2B", "New Markets", "Total",
                         "Net revenues New Markets", round(nr_new[fecha], 2)])

    print(f"  [Canonical/{scen_label}] {len(rows)} filas ({len(set(r[0] for r in rows))} meses)")
    return rows, nr_b2b2c


# ── GSheet manuales ────────────────────────────────────────────────────────────

def _parse_date_sheet(raw):
    """'1/03/2026' → '2026-03-01'"""
    parts = str(raw).strip().split("/")
    if len(parts) != 3:
        return None
    d, m, y = parts
    try:
        return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    except ValueError:
        return None


def _parse_valor_sheet(raw):
    s = str(raw).strip()
    if not s:
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    else:
        try:
            return int(s.replace(".", ""))
        except ValueError:
            try:
                return float(s)
            except ValueError:
                return None


def _norm_periodo(v):
    """datetime | '2026-04' | '2026-04-01' | '1/4/2026' → 'YYYY-MM-01'  (None si no parsea)."""
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return f"{v.year:04d}-{v.month:02d}-01"
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})(?:-\d{1,2})?$", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-01"
    parts = s.replace("-", "/").split("/")
    if len(parts) == 3:
        try:
            a, b, c = (int(x) for x in parts)
            if a > 1000:              # Y/m/d
                y, mo = a, b
            else:                     # d/m/Y
                y, mo = (c + 2000 if c < 100 else c), b
            return f"{y:04d}-{mo:02d}-01"
        except ValueError:
            pass
    return None


def _read_agencias_flat():
    """KR 'Monthly Buying Agencies' (B2B, peso 20) — se carga a mano en
    AGENCIAS_FLAT (agencias_okr.xlsx: periodo | escenario | valor) porque el
    estado de la orden en el datalake muta con el tiempo y hay que congelar el
    snapshot mensual. Vida útil hasta Sep-2026; Oct-2026 cambia la lógica."""
    if not os.path.exists(AGENCIAS_FLAT):
        print(f"  [Agencias] {os.path.basename(AGENCIAS_FLAT)} no existe — KR sin datos")
        return []
    import openpyxl
    wb = openpyxl.load_workbook(AGENCIAS_FLAT, read_only=True, data_only=True)
    ws = wb["agencias"] if "agencias" in wb.sheetnames else wb.active
    rows, skipped = [], 0
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or r[0] is None:
            continue
        periodo   = _norm_periodo(r[0])
        escenario = str(r[1] or "").strip()
        valor_raw = r[2] if len(r) > 2 else None
        if periodo is None or escenario == "" or valor_raw in (None, ""):
            skipped += 1
            continue
        try:
            valor = float(valor_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue
        rows.append([periodo, escenario, "B2B", "Total", "Total",
                     "Monthly Buying Agencies", round(valor, 2)])
    wb.close()
    print(f"  [Agencias] {len(rows)} filas de {os.path.basename(AGENCIAS_FLAT)}"
          + (f" ({skipped} sin valor/incompletas)" if skipped else ""))
    return rows


def _read_manual_krs(token_file):
    """Lee del GSheet Input_OKR los KRs manuales de MANUAL_KRS
    (Sign New Partnership, Air Net Revenue from suppliers)."""
    raw_rows = _read_sheet(token_file)
    SKIP = {"periodo", "escenario", "lob", "kr", "", None}
    rows = []
    skipped = 0
    for r in raw_rows:
        if len(r) < 7:
            skipped += 1
            continue
        periodo_raw, escenario, lob, pais, producto, kr, valor_raw = (
            r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        )
        if str(periodo_raw).strip().lower() in SKIP or str(escenario).strip().lower() in SKIP:
            skipped += 1
            continue
        if kr not in MANUAL_KRS:
            continue
        # Solo el total (Pais=Total, Producto=Total) — si alguien agrega detalle
        # por país/producto en la sheet, no lo sumamos para evitar duplicar.
        if str(pais).strip().lower() != "total" or str(producto).strip().lower() != "total":
            skipped += 1
            continue
        periodo = _parse_date_sheet(periodo_raw)
        valor   = _parse_valor_sheet(valor_raw)
        if periodo is None or valor is None:
            skipped += 1
            continue
        rows.append([periodo, escenario, lob, pais, producto, kr, valor])

    print(f"  [GSheet manual] {len(rows)} filas")
    return rows


# ── Build ──────────────────────────────────────────────────────────────────────

def build(upload=True):
    print("=== okr_builder ===")

    svc        = pnl_common.get_drive_service()
    token_file = os.path.join(DIR, "token_drive.json")

    print("\n--- Descargando fuentes de Drive ---")
    gest_json     = _download_json(svc, GESTIONAL_FILE_ID,   "_pnl_gestional_data.json")
    baseline_data = _download_json(svc, BASELINE_FILE_ID,    "baseline_actuals+projections.json")
    budget_data   = _download_json(svc, BUDGET_FILE_ID,      "budget.json")
    daily_json    = _download_json(svc, DAILY_B2B2C_FILE_ID, "daily_b2b2c_data.json")

    last_actual_ym = gest_json.get("last_actual_ym")
    if not last_actual_ym:
        sys.exit("gestional JSON sin 'last_actual_ym' — no puedo determinar el corte de actuals.")
    print(f"  corte de actuals (gestional.last_actual_ym): {last_actual_ym}")

    print("\n--- Calculando KRs ---")
    rows_baseline, nr_b2b2c_baseline = _compute_from_canonical(baseline_data, "Run Rate/Actuals")
    rows_budget,   nr_b2b2c_budget   = _compute_from_canonical(budget_data,   "Budget")
    rows_hunting  = _compute_b2b2c_hunting_farming(
        daily_json,
        {"Run Rate/Actuals": nr_b2b2c_baseline, "Budget": nr_b2b2c_budget},
        last_actual_ym,
    )
    rows_manual   = _read_manual_krs(token_file)
    rows_agencias = _read_agencias_flat()

    all_rows = rows_baseline + rows_budget + rows_hunting + rows_manual + rows_agencias

    # Resumen
    print(f"\n  Total filas: {len(all_rows):,}")
    by_kr = defaultdict(int)
    for r in all_rows:
        by_kr[r[5]] += 1
    for kr, n in sorted(by_kr.items()):
        print(f"    {kr}: {n}")

    payload = {
        "meta": {
            "concepto":   "okr",
            "gestional_file_id":  GESTIONAL_FILE_ID,
            "baseline_file_id":   BASELINE_FILE_ID,
            "budget_file_id":     BUDGET_FILE_ID,
            "daily_b2b2c_file_id": DAILY_B2B2C_FILE_ID,
            "sheet_id":   SHEET_ID,
            "filas":      len(all_rows),
        },
        "cols": COLS_OUT,
        "rows": all_rows,
    }

    outdir = os.path.join(DIR, "_projections_out")
    os.makedirs(outdir, exist_ok=True)
    local = os.path.join(outdir, OUTPUT_NAME)
    with open(local, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\n  local -> {local}  ({os.path.getsize(local)/1e3:.1f} KB)")

    if upload:
        _upload(local, svc)
    else:
        print("  (--no-upload)")

    return payload


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-upload", action="store_true")
    a = ap.parse_args()
    build(not a.no_upload)
