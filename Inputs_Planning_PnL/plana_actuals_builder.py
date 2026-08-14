"""
plana_actuals_builder.py
Genera la PLANA de ACTUALS por AÑO FISCAL (Abr–Mar) desde los '00 - Actuals YYYY - Plana Python.xlsx'.
Un FY combina dos archivos calendario: Abr-Dic del (fy-1) + Ene-Mar del (fy).
Ej: FY2027 = Abr-Dic 2026 (file 2026) + Ene-Mar 2027 (file 2027).

Uso:
    python plana_actuals_builder.py <fy> [--con-ppa]
    <fy> = año fiscal (2024, 2025, 2026, 2027)

Default = SIN PPA -> suma Reverso AxI por período (RunRate para Oct25-Mar26, Budget para Abr26-Mar27).
Output: actuals_fy<YY>_abr<fy-1>_mar<fy>.csv  (mismo formato que las planas de proyecciones)
"""
import os
import sys
import pandas as pd
import pnl_common

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import config
from pnl_common import (
    DIM_COLS, OUT_COLS, LOBS_KEEP, MMM,
    homolog, to_num, lower_all,
    load_glosario, add_axi, load_axi_sheet,
)

BITUBIA      = os.path.join(pnl_common.get_base_dir(), "Proyectos IA", "BITUBIA")
AXI_RR_DATES = config.AXI_RR_DATES

NIVELES = ["P&L N1", "P&L N2", "P&L N3", "P&L N4", "P&L N5", "P&L N6", "P&L Managerial View"]

# ── Exclusiones Actuals (Parte 2.3) ──────────────────────────────────────────
EXCL_FUTURO = {"dfinance", "intercompany", "koin"}
EXCL_LINEA_ACT = {
    "(-)fx intercompany", "(-)other operating income", "(+)fx intercompany", "fx intercompany",
    "fx/other", "income tax", "interest expense", "interest income",
    "net income (loss) attributable to non controlling interest",
    "other comprehensive income (loss) attributable to despegar group",
    "other operating income", "other operating income (expense)", "withholdings tax",
    "- cancelled gb", "cancelled gb", "leases interest",
}


# ── Lectura de un archivo calendario ─────────────────────────────────────────
def read_actuals_file(year, wanted_dates, bitubia_dir=None):
    bitubia = bitubia_dir or BITUBIA
    path = os.path.join(bitubia, f"00 - Actuals {year} - Plana Python.xlsx")
    if not os.path.exists(path):
        print(f"  (falta {os.path.basename(path)} -> se omite ese tramo)")
        return [], []
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = None
    for sn in wb.sheetnames:
        if any(k in sn.strip().lower() for k in ("ajuste", "adj")):
            continue
        ws = wb[sn]
        hdr = next(ws.iter_rows(values_only=True), None)
        if not hdr:
            continue
        hl = [str(h).strip().lower() for h in hdr]
        if "futuronombre" in hl and "linea p&l" in hl:
            sheet = sn
            break
    if sheet is None:
        wb.close()
        raise RuntimeError(f"No encontré hoja de datos (FuturoNombre + Linea P&L) en {path}")
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    raw_hdr = list(next(it))
    hdr = [str(h).strip() if h is not None else "" for h in raw_hdr]
    hl = [h.lower() for h in hdr]
    ci = {k: hl.index(k.lower()) for k in ["FuturoNombre", "LoB", "CodigoPais", "Prod_Corregido", "Linea P&L"]}

    import datetime as _dt
    def _norm(h):
        if isinstance(h, _dt.datetime):
            return f"{h.year:04d}-{h.month:02d}-01"
        return str(h).strip()[:10]
    date_idx = {j: _norm(h) for j, h in enumerate(raw_hdr) if _norm(h) in wanted_dates}
    recs = []
    for r in it:
        rec = {
            "FuturoNombre": r[ci["FuturoNombre"]],
            "LoB":          r[ci["LoB"]],
            "CodigoPais":   r[ci["CodigoPais"]],
            "Prod_Corregido": r[ci["Prod_Corregido"]],
            "Linea P&L":    r[ci["Linea P&L"]],
        }
        for j, d in date_idx.items():
            rec[d] = to_num(r[j]) if j < len(r) else None
        recs.append(rec)
    wb.close()
    print(f"  {os.path.basename(path)} [{sheet}]: {len(recs):,} filas, meses={sorted(date_idx.values())}")
    return recs, sorted(date_idx.values())


# ── Homologación Actuals (Parte 2.4) ─────────────────────────────────────────
def homologate_actuals(df, g):
    df = df.copy()
    df["Marca"]    = df["FuturoNombre"].apply(lambda x: homolog(g["marca"], x, 0, x))
    df["Pais"]     = df["CodigoPais"].apply(lambda x: homolog(g["paises"], x, 0, x))
    df["Producto"] = df["Prod_Corregido"].apply(lambda x: homolog(g["producto"], x, 0, x))
    df["LoB_h"]    = df["LoB"].apply(lambda x: homolog(g["lob"], x, 0, None))
    df["Canal"]    = df["LoB"].apply(lambda x: homolog(g["lob"], x, 1, None))
    miss = df["LoB_h"].isna()
    if miss.any():
        sp = df.loc[miss, "LoB"].astype(str).str.split("-", n=1, expand=True)
        df.loc[miss, "LoB_h"] = sp[0]
        df.loc[miss, "Canal"] = sp[1] if sp.shape[1] > 1 else "Sin Canal"
    for i, col in enumerate(NIVELES):
        df[col] = df["Linea P&L"].apply(lambda x, i=i: homolog(g["linea"], x, i, x))
    df = df.drop(columns=["LoB"]).rename(columns={"LoB_h": "LoB"})
    return df


# ── AXI actuals (Parte 4) ────────────────────────────────────────────────────
def load_axi_actuals(g):
    rr = load_axi_sheet(g, "RunRate", AXI_RR_DATES)   # Oct25-Mar26
    bg = load_axi_sheet(g, "Budget")                  # Abr26-Mar27 (keep_dates=None = todas)
    return pd.concat([rr, bg], ignore_index=True)


# ── Build ────────────────────────────────────────────────────────────────────
def build(fy, con_ppa=False, bitubia_dir=None):
    print(f"=== plana actuals FY{fy} ({'CON' if con_ppa else 'SIN'} PPA) ===")
    prev_dates = [f"{fy-1}-{m:02d}-01" for m in range(4, 13)]   # Abr-Dic (fy-1)
    cur_dates  = [f"{fy}-{m:02d}-01" for m in (1, 2, 3)]        # Ene-Mar (fy)
    r1, d1 = read_actuals_file(fy - 1, set(prev_dates), bitubia_dir)
    r2, d2 = read_actuals_file(fy,     set(cur_dates),  bitubia_dir)
    recs = r1 + r2
    all_dates = d1 + d2
    if not recs:
        sys.exit("No hay datos para ese FY.")
    df = pd.DataFrame(recs)

    # 2.3 exclusiones
    df = df[~df["FuturoNombre"].astype(str).str.strip().str.lower().isin(EXCL_FUTURO)]
    df = df[~df["Linea P&L"].astype(str).str.strip().str.lower().isin(EXCL_LINEA_ACT)]

    g = load_glosario()
    df = homologate_actuals(df, g)

    # 2.5-2.7 melt + groupby + filtro
    long = df.melt(id_vars=["Marca", "LoB", "Canal", "Pais", "Producto"] + NIVELES,
                   value_vars=all_dates, var_name="Fecha", value_name="Monto USD")
    long["Monto USD"] = pd.to_numeric(long["Monto USD"], errors="coerce")
    long = long.dropna(subset=["Monto USD"])
    long = long[long["Monto USD"] != 0]
    plana = long.groupby(DIM_COLS + ["Fecha"], as_index=False, dropna=False)["Monto USD"].sum()
    plana = lower_all(plana)[OUT_COLS]
    plana = plana[plana["LoB"].isin(LOBS_KEEP)]

    if not con_ppa:
        axi = load_axi_actuals(g)
        axi = axi[axi["LoB"].isin(LOBS_KEEP)]
        base_dates = set(plana["Fecha"].unique())
        axi = axi[axi["Fecha"].isin(base_dates)]
        print(f"  reverso AxI (por período): {len(axi):,} filas a sumar (meses cerrados: {len(base_dates)})")
        plana = add_axi(plana, axi)[OUT_COLS]

    yy = str(fy)[-2:]
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       f"actuals_fy{yy}_abr{fy-1}_mar{fy}.csv")
    plana.to_csv(out, index=False, encoding="utf-8")
    print(f"\nOK -> {out}")
    return plana


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    con_ppa = "--con-ppa" in sys.argv
    if not args:
        sys.exit("Uso: python plana_actuals_builder.py <fy> [--con-ppa]   (ej: 2027)")
    build(int(args[0]), con_ppa)
